import datetime
import openpyxl
import openpyxl.styles
import openpyxl.cell
from openpyxl.styles import PatternFill
from openpyxl.worksheet.properties import WorksheetProperties
import openpyxl.worksheet.dimensions


ID = 0
NAME = 1
NUMBER_OF_CARDS = 2


BANK = 0
DATE = 1
TIME = 2
CARD = 3
OPERATION = 4
MONEY = 5
BALANCE = 6
CURRENCY = 7

months = {"01": 'January', "02": 'February', "03": 'March', "04": 'April', "05": 'May', "06": 'June', "07": 'July',
          "08": 'August', "09": 'September', "10": 'October', "11": 'November', "12": 'December'}


def file_to_banks():
    banks1 = []
    text_file = open("./banks.txt").read()
    lines = text_file.split('\n')
    length = len(lines)
    i = 0
    while i < length - 1:
        for line in lines:
            list1 = []
            index1 = line.index(",")
            a = int(line[:index1])
            line_a = line[index1 + 2:]
            index2 = line_a.index(",")
            b = line[index1 + 2:index2 + index1 + 2]
            c = line_a[index2 + 2:-1]
            list1.append(a)
            list1.append(b)
            list1.append(c)
            banks1.append(list1)
            i += 1
    return banks1


def file_to_operations():
    operations1 = []
    text_file = open("./operations.txt").read()
    lines = text_file.split('\n')
    length = len(lines)
    i = 0
    while i < length - 1:
        for line in lines:
            d = ''
            e = ''
            f = ''
            g = ''
            list1 = []
            index1 = line.index(",")
            a = line[:index1]
            line_a = line[index1+2:]
            index2 = line_a.index(",")
            b = line_a[:index2]
            line_b = line_a[index2+2:]
            index3 = line_b.index(",")
            c = line_b[:index3]
            letter = line_b[index3+2:index3+3]
            line_c = line_b[index3+2:]
            if letter != "*":
                index4 = line_c.index(":")
                e = line_c[:index4]
                line_d = line_c[index4+2:]
                index5 = line_d.index("*")
                index6 = line_d.index(":")
                d = line_d[index5:index6]
                line_e = line_d[index6+2:]
                index7 = line_e.index(" ")
                f = line_e[:index7]
                index8 = line_e.index(":")
                g = line_e[index8+2:-2-len("CUR")]
            if letter == "*":
                index4 = line_c.index(",")
                d = line_c[:index4]
                line_d = line_c[index4+2:]
                e = line_d[:1]
                index5 = line_d.index(" ")
                f = line_d[1:index5]
                index7 = line_d.index(":")
                line_e = line_d[index7+2:]
                g = line_e[:-2-len("CUR")]
            h = line[-1-len("CUR"):-1]
            list1.append(a)
            list1.append(b)
            list1.append(c)
            list1.append(d)
            list1.append(e)
            list1.append(f)
            list1.append(g)
            list1.append(h)
            operations1.append(list1)
            i += 1
    return operations1


banks = file_to_banks()
operations = file_to_operations()


def sort_by_date():
    d = []
    i = 0
    while i < len(operations):
        d.append(operations[i][TIME]+" "+operations[i][DATE])
        i += 1
    return sorted(d, key=lambda x: datetime.datetime.strptime(x, '%H:%M %d-%m-%Y'))


def date_to_str(month_year):
    usual_month = "XX"
    index = len(usual_month)
    month = month_year[:index]
    return months[month] + " " + month_year[index+1:]


date = sort_by_date()


def sorted_operations(date_sorted_by_date):
    list2 = []
    i = 0
    while i < len(date_sorted_by_date):
        j = 0
        while j < len(operations):
            if date_sorted_by_date[i] == operations[j][TIME] + " " + operations[j][DATE]:
                list2.append(operations[j])
            j += 1
        i += 1
    return list2


operations2 = sorted_operations(date)


def number_by_card(card):
    i = 0
    number_of_bank = 0
    while i < len(operations):
        if card == operations[i][CARD]:
            number_of_bank = int(operations[i][BANK])
        i += 1
    return number_of_bank


def name_by_number(number):
    i = 0
    while i < len(banks):
        if banks[i][ID] == number:
            return banks[i][NAME]
        i += 1


def names_of_banks():
    bank_names = []
    i = 0
    while i < len(banks):
        bank_names.append(banks[i][ID])
        i += 1
    return bank_names


def all_cards_in_banks(date_r):
    all_cards = []
    names = names_of_banks()
    i = len(date_r)
    while i > 0:
        one_card = date_r[i-1][CARD]
        bank = int(date_r[i-1][BANK])
        if one_card in all_cards:
            i -= 1
            continue
        elif one_card not in all_cards and bank in names:
            all_cards.append(one_card)
            i -= 1
        else:
            i -= 1
    return all_cards


cards = all_cards_in_banks(operations2)


def balance(card):
    balance_of_card = []
    i = 0
    while i < len(cards):
        if cards[i] == card:
            j = len(operations2)
            while j > 0:
                if operations2[j-1][CARD] == card:
                    balance_of_card.append(operations2[j-1][BALANCE])
                j -= 1
        i += 1
    return balance_of_card


def balance_in_month(month):
    operations_in_month = []
    i = 0
    while i < len(operations2):
        current_date = operations2[i][DATE][3:]
        current_card = operations2[i][CARD]
        if current_date == month and current_card in cards:
            operations_in_month.append(operations2[i])
        i += 1
    j = 0
    cards_in_month = []
    while j < len(operations_in_month):
        if operations_in_month[j][CARD] not in cards_in_month:
            cards_in_month.append(operations_in_month[j][CARD])
        j += 1
    report_for_card = []
    p = 0
    while p < len(cards_in_month):
        j = 0
        while j < len(operations_in_month):
            list_of_operations = []
            if cards_in_month[p] == operations_in_month[j][CARD]:
                list_of_operations.append(operations_in_month[j][CARD])
                a = operations_in_month[j][OPERATION]
                b = operations_in_month[j][MONEY]
                if a == "+" or a == "Transfer":
                    list_of_operations.append(b)
                    list_of_operations.append("")
                elif a == "-" or a == "Withdrawal":
                    list_of_operations.append("")
                    list_of_operations.append(b)
                report_for_card.append(list_of_operations)
            j += 1
        p += 1
    return report_for_card, operations_in_month


def received_spent_delta(some_list):
    length = len(some_list)
    another_list = []
    for card in cards:
        received = 0
        spent = 0
        delta = 0
        i = 0
        while i < length:
            if card == some_list[i][0]:
                if len(some_list[i][1]) != 0:
                    received += float(some_list[i][1])
                elif len(some_list[i][2]) != 0:
                    spent += float(some_list[i][2])
                delta = received - spent
            i += 1
        if received != 0 or spent != 0:
            another_list.append([card, received, spent, delta])
    return another_list, length


def rsd_for_all_cards(some_list):
    another_list = []
    plus = 0
    minus = 0
    i = 0
    while i < len(some_list):
        plus += float(some_list[i][1])
        minus += float(some_list[i][2])
        i += 1
    another_list.append(plus)
    another_list.append(minus)
    another_list.append(plus-minus)
    return another_list


def rsd_to_operations_for_report(month):
    z, operations_in_month = balance_in_month(month)
    another_data = []
    i = 0
    while i < len(operations_in_month):
        cur_bank = name_by_number(int(operations_in_month[i][BANK]))
        cur_date = operations_in_month[i][DATE]
        cur_time = operations_in_month[i][TIME]
        cur_card = operations_in_month[i][CARD]
        cur_operation = operations_in_month[i][OPERATION] + " " + operations_in_month[i][MONEY]
        cur_operation += " " + operations_in_month[i][CURRENCY]
        cur_balance = operations_in_month[i][BALANCE] + " " + operations_in_month[i][CURRENCY]
        another_data.append([cur_bank, cur_date, cur_time, cur_card, cur_operation, cur_balance])
        i += 1
    return another_data


def the_last_func(some_data):
    i = 0
    while i < len(some_data):
        if some_data[i][OPERATION][:10] == "Withdrawal":
            operation = some_data[i][OPERATION][11:]
            some_data[i][OPERATION] = "-" + operation
        elif some_data[i][OPERATION][:8] == "Transfer":
            operation = some_data[i][OPERATION][9:]
            some_data[i][OPERATION] = "+" + operation
        i += 1
    return some_data


def export_to_excel(chosen_date, card, res_spe_del):
    its_date = chosen_date.replace('-', "_")
    rsd = rsd_to_operations_for_report(chosen_date)
    data = the_last_func(rsd)
    pink_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
    wb = openpyxl.Workbook()
    ws1 = wb.active
    openpyxl.worksheet.dimensions.ColumnDimension(ws1, auto_size=True)
    ws1.title = "report"
    ws1.cell(1, 1, "Bank").fill = pink_fill
    ws1.cell(1, 2, "Date").fill = pink_fill
    ws1.cell(1, 3, "Time").fill = pink_fill
    ws1.cell(1, 4, "Card").fill = pink_fill
    ws1.cell(1, 5, "Operation").fill = pink_fill
    ws1.cell(1, 6, "Balance").fill = pink_fill
    i = 0
    if card == "all cards":
        while i < len(data):
            j = 0
            while j < len(data[i]):
                ws1.cell(i+2, j+1, data[i][j])
                j += 1
            i += 1
        z = i + 4
        ws1.cell(z, 1, "Bank").fill = pink_fill
        ws1.cell(z, 2, "Month").fill = pink_fill
        ws1.cell(z, 3, "Card").fill = pink_fill
        ws1.cell(z, 4, "Received").fill = pink_fill
        ws1.cell(z, 5, "Spent").fill = pink_fill
        ws1.cell(z, 6, "Delta").fill = pink_fill
        z += 1
        k = 0
        while k < len(res_spe_del):
            j = 0
            while j < len(res_spe_del[k]):
                bank_name = name_by_number(number_by_card(res_spe_del[k][0]))
                ws1.cell(z, 1, bank_name)
                ws1.cell(z, 2, chosen_date)
                count = 3
                while count < 7:
                    ws1.cell(z, count, res_spe_del[k][j])
                    j += 1
                    count += 1
                j += 1
            k += 1
            z += 1
    else:
        row = 2
        while i < len(data):
            if card == data[i][CARD]:
                column = 1
                j = 0
                while j < len(data[i]):
                    ws1.cell(row, column, data[i][j])
                    column += 1
                    j += 1
                row += 1
            i += 1
        z = i + 4
        ws1.cell(z, 1, "Bank").fill = pink_fill
        ws1.cell(z, 2, "Month").fill = pink_fill
        ws1.cell(z, 3, "Card").fill = pink_fill
        ws1.cell(z, 4, "Received").fill = pink_fill
        ws1.cell(z, 5, "Spent").fill = pink_fill
        ws1.cell(z, 6, "Delta").fill = pink_fill
        z += 1
        k = 0
        while k < len(res_spe_del):
            j = 0
            while j < len(res_spe_del[k]):
                if card == res_spe_del[k][0]:
                    bank_name = name_by_number(number_by_card(res_spe_del[k][0]))
                    ws1.cell(z, 1, bank_name)
                    ws1.cell(z, 2, chosen_date)
                    count = 3
                    while count < 7:
                        ws1.cell(z, count, res_spe_del[k][j])
                        count += 1
                        j += 1
                j += 1
            k += 1
    wb.save(filename=f'Report_{its_date} for {card}.xlsx')
    return card
